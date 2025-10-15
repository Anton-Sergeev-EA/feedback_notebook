import os  # Импортируем модуль для работы с файловой системой.
from openpyxl import load_workbook  # Импортируем функцию для работы с Excel файлами.
from tkinter import messagebox  # Импортируем модуль для всплывающих окон.


def check_company(company_name: str) -> bool:
    """
    Проверяет наличие компании в базе данных.
    :param company_name: название компании для проверки.
    :return: True, если компания существует, False — если нет.
    """
    try:
        # Проверяем существование файла.
        if not os.path.exists('applications.xlsx'):
            return False
        
        # Открываем файл Excel.
        wb = load_workbook('applications.xlsx')
        sheet = wb.active
        
        # Проходим по строкам, начиная со второй.
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Проверяем наличие данных в строке.
            if len(row) > 1 and row[1] == company_name:
                return True
        
        return False
    
    except Exception as e:
        messagebox.showerror(
            "Ошибка",
            f"Ошибка при проверке компании: {str(e)}"
        )
        return False
