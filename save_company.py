import os # Импортируем модуль для работы с файловой системой.
from openpyxl import Workbook, load_workbook # Импортируем классы для работы
# с Excel.
from tkinter import messagebox # Импортируем модуль для всплывающих окон.


def save_company(company_name: str, position: str) -> bool:
    """
    Сохраняет данные о компании в Excel-файл.
    :param company_name: название компании.
    :param position: должность.
    :return: True при успешном сохранении, False при ошибке.
    """
    try:
        # Проверяем существование файла перед открытием.
        if not os.path.exists('applications.xlsx'):
            # Создаем новый файл, если его нет.
            wb = Workbook()
            sheet = wb.active
            
            # Добавляем заголовки столбцов.
            sheet.cell(row=1, column=1, value='ID')
            sheet.cell(row=1, column=2, value='Название компании')
            sheet.cell(row=1, column=3, value='Должность')
        else:
            wb = load_workbook('applications.xlsx')
        
        # Получаем активный лист.
        sheet = wb.active
        
        # Находим последнюю заполненную строку.
        last_row = 2  # Начинаем поиск со второй строки.
        while sheet.cell(row=last_row, column=2).value is not None:
            last_row += 1
        
        # Заполняем новые данные.
        company_id = last_row - 1
        sheet.cell(row=last_row, column=1, value=company_id)
        sheet.cell(row=last_row, column=2, value=company_name)
        sheet.cell(row=last_row, column=3, value=position)
        
        # Сохраняем изменения.
        wb.save('applications.xlsx')
        return True
    
    except PermissionError:
        messagebox.showerror(
            "Ошибка",
            "Файл открыт в другой программе. Закройте его и повторите попытку."
        )
        return False
    
    except Exception as e:
        messagebox.showerror(
            "Ошибка",
            f"Ошибка при сохранении компании: {str(e)}"
        )
        return False
